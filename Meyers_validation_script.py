from openpyxl import load_workbook
import pandas
import validators
import re
import sys
import traceback

try:
    fileName = str(input('\n Please enter Meyers output file name(.xlsx) : '))
    try:
        excel_data_df = pandas.read_excel(fileName, sheet_name='Scraped_Output')
        wb=load_workbook(fileName)

        ws=wb.worksheets[0]

    
        ws['I1']='Status'
        ws['J1']='Comments'
        
   
        temp=['Project_URL','Project_Name','Builder_Name','Project_Address','Project_Phone_Number','Latitude','Longitude','Project_Status']
        Temp_Chck=excel_data_df.columns.ravel()

       
        Unmatch=[]
        for i in range(len(temp)):
                if temp[i] not in Temp_Chck:
                    Unmatch.append(temp[i] + '-Row :'+ str(i+1) + ' Templete is Missing' )
                    ws['K1']='Template Check Fail'
                    wb.save(fileName)
                    print('\n Error: Please Check Template!')
                    #sys.exit()
                else:
                    pass
                   

        Final_Unmatch=[]
        if Unmatch:
            pass
        else:
            ws['K1']='Template Check Sucess'
            

        
        #________Url Format Check__________________________________________________________

        Ex_Lat=list(excel_data_df.iloc[1,:].values)

        Url_data=excel_data_df.iloc[:,0:1].values
        Total_Url=[]
        url=[]
        for i in Url_data:
            for j in i:
                Total_Url.append(j)
                valid=validators.url(str(j))
                if valid==True:
                    pass
                else:
                    url.append(j)
        line=[]
        content=[]
        c1=[]
        for UU in range(len(Total_Url)):
            c1.append(UU)
            if Total_Url[UU] in url:
                line.append(UU+2)
                content.append('Please Check URL')
            else:
                pass

        res = dict(zip(line, content))




        #_______________Phone Number_____________________________________________

        Line_Num=[]
        Er_Con=[]
        Url_data=excel_data_df.iloc[:,0:1].values
        num=Url_data=excel_data_df.iloc[:,4].values
        for i in range(len(Url_data)):
            pattern = r'[-()|+.]'
            mod_string = re.sub(pattern, '',str(num[i]).replace(' ',''))
            if mod_string.isnumeric() == True:
                if len(str(mod_string)) ==10:
                    pass
                elif len(str(mod_string)) == 20:
                        if (mod_string[0:10]) != (mod_string[10:21]):
                            pass
                        else:
                            Line_Num.append(i+2)
                            Er_Con.append('Same Phone Number Repeated Twice Please Check')
                elif len(str(mod_string)) ==30:
                        if (mod_string[0:10]) != (mod_string[10:20]) != (mod_string[20:30]):
                            pass
                        else:
                            Line_Num.append(i+2)
                            Er_Con.append('Same Phone Number Repeated Thrice Please Check')
                else:
                    Line_Num.append(i+2)
                    Er_Con.append('Please Check Phone Number length')
            elif mod_string == 'nan':
                    pass         
            else:
                Line_Num.append(i+2)
                Er_Con.append('Please Check Phone Number not in Numeric')

        res1 = dict(zip(Line_Num, Er_Con))
        for key in res:
            if key in res1:
                va=str(res[key]) + '| ' + str(res1[key])
                res.update({key:va})
            else:
                pass
        res1.update(res)


        #Junk Processs__________________________
        spcl=[]
        spclCharsContent = open('special_characters.txt', 'r', encoding='utf-8').read()
        spclChars = spclCharsContent.strip().split(' ')
        for spclChar in spclChars:
            spcl.append(spclChar)

        Url_data=excel_data_df.iloc[:,0:1].values
        junk=[]
        junk_line=[]
        w=0
        while w!=len(Url_data):
            c=0
            n=list(excel_data_df.iloc[w,:].values)
            for i in range(len(spcl)):
                for j in range(len(n)):
                    if spcl[i] not in str(n[j]):
                        pass
                    else:
                        junk_line.append(w+2)
                        junk.append('Please Check the Entire Record,Junk is Available')
                
            w+=1

        res2 = dict(zip(junk_line,junk))
        for key in res2:
            if key in res1:
                ju=str(res1[key]) + ' | ' +str(res2[key])
                res1.update({key:ju})
            else:
                pass

        res2.update(res1)



        #_Latitude_____________________________
        lati_line=[]
        lati=[]
        Url_data=excel_data_df.iloc[:,0:1].values
        num=Url_data=excel_data_df.iloc[:,5].values
        for i in range(len(Url_data)):
            pattern = r'[.]'
            mod_string = re.sub(pattern, '',str(num[i]).replace(' ',''))
            if mod_string.isnumeric() == False:
                if '-' in str(num[i]):
                    lati_line.append(i+2)
                    lati.append('Negative value in Latitude')
                elif str(num[i]).isalpha() == False:
                    lati_line.append(i+2)
                    lati.append('Please Check Latitude not in Numeric')
                else:
                    pass
            else:
                if "." not in str(num[i]):
                    lati_line.append(i+2)
                    lati.append('Dot missing in Latitude')
                elif str(num[i])[-2]=='.':
                    lati_line.append(i+2)
                    lati.append('Dot missing in Latitude')
                else:
                    pass
                
                    

        res3 = dict(zip(lati_line,lati))


        for key in res3:
            if key in res2:
                la=str(res2[key]) + ' | ' + str(res3[key])
                res2.update({key:la})
            else:
                pass

        res3.update(res2)


        #______Longitude___________________________________
        lon_Line=[]
        lon=[]
        Url_data=excel_data_df.iloc[:,0:1].values
        num=Url_data=excel_data_df.iloc[:,6].values
        for i in range(len(Url_data)):
            pattern = r'[.-]'
            mod_string = re.sub(pattern, '',str(num[i]).replace(' ',''))
            if str(num[i]) == "nan":
                    pass
            elif mod_string.isnumeric()==False:
                    lon_Line.append(i+2)
                    lon.append('Please Check Longitude not in Numeric')
            else:
                if  '.' not in str(num[i]) and '-' not in str(num[i]):
                    lon_Line.append(i+2)
                    lon.append('Positive Number in Longitude,Dot missing')

                elif str(num[i])[-2]=='.':
                    lon_Line.append(i+2)
                    lon.append('Dot missing in Longitude')
                    
                elif '.' not in str(num[i]):
                    lon_Line.append(i+2)
                    lon.append('Dot missing in Longitude')


                elif '-' not in str(num[i]):
                    lon_Line.append(i+2)
                    lon.append('Positive Number in Longitude')

                else:
                    pass

                        

        res4 = dict(zip(lon_Line,lon))

        for key in res4:
            if key in res3:
                lo=str(res3[key]) + ' | ' + str(res4[key])
                res3.update({key:lo})
            else:
                pass
        res4.update(res3)



        #___Project_Status_____________________________________________________________

        status_list=['Sold Out','Now Open','Final Opportunity','Coming Soon','Now Selling','Limited Availability','CLOSE OUT','COMING Summer 2022','COMING Spring 2022','Phase 2 Coming Soon!','Homes coming soon!','Almost Sold Out',
                    'New Phase! Limited Release!','Coming Fall of 2022!','Model Open, Now Selling!','Active','Model Now Open!','Final Phase Coming Soon','Next Phase Coming Soon','New Phase Available','Coming 2023','NOW SELLING PHASE 3!',
                    'Model Grand Opening!','Final Homes Remaining','closeout','Models Coming Soon','SOLD OUT!','Selling out','final phase now open','Almost Sold Out','Model for Sale','Now Pre-Selling','New Model Now Open','New Phase Coming Soon',
                    'CLOSE OUT','NOW SELLING!','NOW SELLING','COMING SOON','MODEL GRAND OPENING','GRAND OPENING','COMING SOON!','NOW OPEN!','NOW OPEN','Now Open!','Closeout!','Coming Soon!','Grand opening!','NOW SELLING','Final Phase Coming!','NEW PHASE COMING',
                    'FINAL OPPORTUNITIES','NEW PHASE COMING','Sold Out!','Now Selling!','NOW PRE-SELLING!','New Section Coming Soon!','Model Home Now Open','New Phase Now Open','New Section Now Open!','Temporarily Sold Out','New Models Now Open!','Now Selling Final Phase!',
                    'FINAL OPPORTUNITIES!!!','COMING SOON!!!','COMING SOON!!!','COMING SOON!!!','COMING SOON!!!','NOW SELLING!!!!','CURRENTLY SOLD OUT - NEW PHASE COMING','GRAND OPENING!!!','NOW SELLING PHASE 3!!!','NEW PHASE COMING!!!','NEW PHASE COMING SOON!!! NOW PRE-SELLING!!!',
                    'NOW PRE-SELLING','NOW PRESELLING','FINAL OPPORTUNITIES REMAINING','MODEL NOW OPEN','SELLING FAST','PHASE TWO NOW SELLING','MODEL NOW OPEN','FINAL HOMES','SOLD OUT','NOW SELLING PHASE II','NEW PHASE COMING SOON','FINAL HOME','Coming Soon ','Final Opportunities!',
                    'New Phase - Now Selling!','Quick-Move Homes Available Now!','Limited Availability','Now selling','Grand Opening!','Final Opportunities!','Final Opportunities','Model now open','Grand Opening','Final Opportunities!','coming soon!']



        Url_data=excel_data_df.iloc[:,0:1].values
        status_line=[]
        status=[]

        sta=Url_data=excel_data_df.iloc[:,7:].values
        for i in range(len(Url_data)):
                    if str(sta[i])=='[nan]':
                        pass
                    elif sta[i] not in status_list:
                            status_line.append(i+2)
                            status.append('Please Check Status')
                    else:
                        pass


        res5=dict(zip(status_line,status))

        for key in res5:
            if key in res4:
                sta=str(res4[key]) + ' | ' + str(res5[key])
                res4.update({key:sta})
            else:
                pass
        res5.update(res4)


    #_____________Address____________________________


        Url_data=excel_data_df.iloc[:,0:1].values
        Add=Url_data=excel_data_df.iloc[:,3].values
        Add_line=[]
        Address=[]
        for add in range(len(Url_data)):
            street_address_validate_pattern = "^(\\d{1,}) [a-zA-Z0-9\\s]+(\\,)? [a-zA-Z]+(\\,)? [A-Z]{2} [0-9]{5}$"
            try:
                if (re.match(street_address_validate_pattern,Add[add])) == None:
                    Add_line.append(add+2)
                    Address.append('Please Check Address')
            except TypeError:
                pass
        res6=dict(zip(Add_line,Address))

        for key in res6:
            if key in res5:
                ad=str(res5[key]) + ' | ' +str(res6[key])
                res5.update({key:ad})
            else:
                pass
        
        res6.update(res5)


    #______________duplicate Check______________

        Url_data=excel_data_df.iloc[:,0:1].values
        Dub=[]
        Dub_Line=[]
        Main_line=[]
        w=0
        while w!=len(Url_data):
            n=str(excel_data_df.iloc[w,:].values)
            if n not in Main_line:
                Main_line.append(n)
            else:
                Dub.append('It is a Duplicate Record')
                Dub_Line.append(w+2)
                
            w+=1

        res7=dict(zip(Dub_Line,Dub))


        for key in res7:
            if key in res6:
                du=str(res6[key]) + ' | ' + str(res7[key])
                res6.update({key:du})
            else:
                pass
        
        
        res7.update(res6)

        for fin in res7.keys():
            ws['I'+str(fin)] ='Error'
            ws['J'+str(fin)] =res7[fin]

        wb.save(fileName)
  

    except FileNotFoundError:
        print('\n Error: Please enter valid file name!')
        sys.exit()

except Exception:
		print(traceback.format_exc())

